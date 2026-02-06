"""
Generate manager report with Weekly Hours and Opportunities sheets.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.config import get_settings


# Map detailed categories to simplified manager categories
MANAGER_CATEGORY_MAP = {
    'Travel': 'Travel',
    'Prep - Demo/ Presentation': 'Prep for Customer',
    'Customer - Demo/ Presentation': 'Customer Demo',
    'Discovery': 'Customer Demo',
    'POC': 'Customer Demo',
    'Internal Meeting': 'Internal Meeting',
    'RFI/RFP/RFQ': 'RFP/RFI/RFQ',
    'Admin': 'Admin',
    'Support': 'Admin',
    'Training': 'Training',
    'Time Off': 'Time Off',
}

# Column order for Weekly Hours sheet
WEEKLY_HOURS_COLUMNS = [
    'Week of',
    'Total Hours',
    'Travel',
    'Prep for Customer',
    'Customer Demo',
    'Internal Meeting',
    'RFP/RFI/RFQ',
    'Admin',
    'Training',
    'Time Off',
]


def load_project_codes_full() -> pd.DataFrame:
    """Load project codes with all columns from Excel."""
    settings = get_settings()
    path = Path(settings['paths']['project_codes'])

    df = pd.read_excel(path)

    # Detect format and normalize
    if 'JDA OpptyID' in df.columns:
        # New format - keep original columns, add normalized ones
        df['code'] = df['JDA OpptyID']
        df['company'] = df['Account Name']
    elif 'Project Code' in df.columns:
        # Old format with headers
        df['code'] = df['Project Code']
        df['company'] = df['Company']
        # Create placeholder columns for report (old format doesn't have these)
        df['JDA OpptyID'] = df['code']
        df['Account Name'] = df['company']
        df['Opportunity Name'] = df.get('Project Description', '')
        for col in ['JDA Industry', 'Stage', 'Booking Amount [USD]', 'Close Date', 'Next Step']:
            if col not in df.columns:
                df[col] = ''
    else:
        # Legacy format (positional columns)
        df.columns = ['company', 'description', 'code']
        df['JDA OpptyID'] = df['code']
        df['Account Name'] = df['company']
        df['Opportunity Name'] = df['description']
        for col in ['JDA Industry', 'Stage', 'Booking Amount [USD]', 'Close Date', 'Next Step']:
            df[col] = ''

    df['company_lower'] = df['company'].str.lower().str.strip()

    return df


def get_weeks_back(weeks_back: int) -> list[str]:
    """Get list of week_beginning dates for last N weeks."""
    today = datetime.now()
    # Find most recent Sunday
    days_since_sunday = (today.weekday() + 1) % 7
    current_sunday = today - timedelta(days=days_since_sunday)

    weeks = []
    for i in range(weeks_back):
        week_date = current_sunday - timedelta(weeks=i)
        weeks.append(week_date.strftime("%Y-%m-%d"))

    return list(reversed(weeks))  # Oldest first


def generate_weekly_hours_df(df: pd.DataFrame, weeks_back: int) -> pd.DataFrame:
    """Generate Weekly Hours pivot table."""
    # Filter out summary rows
    df = df[df["category"] != ">>> WEEK TOTAL"].copy()

    # Map categories to manager categories
    df["manager_category"] = df["category"].map(MANAGER_CATEGORY_MAP)

    # Filter to last N weeks
    weeks = get_weeks_back(weeks_back)
    df = df[df["week_beginning"].isin(weeks)]

    # Pivot table: rows = weeks, columns = manager categories
    pivot = df.pivot_table(
        index="week_beginning",
        columns="manager_category",
        values="hours",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    # Calculate total hours per week
    category_cols = [c for c in pivot.columns if c != "week_beginning"]
    pivot["Total Hours"] = pivot[category_cols].sum(axis=1)

    # Rename week column
    pivot = pivot.rename(columns={"week_beginning": "Week of"})

    # Ensure all columns exist (some categories might not have hours)
    for col in WEEKLY_HOURS_COLUMNS:
        if col not in pivot.columns:
            pivot[col] = 0

    # Reorder columns
    pivot = pivot[WEEKLY_HOURS_COLUMNS]

    # Sort by week (oldest first)
    pivot = pivot.sort_values("Week of")

    # Add TOTAL row
    total_row = {"Week of": "TOTAL"}
    for col in WEEKLY_HOURS_COLUMNS[1:]:
        total_row[col] = pivot[col].sum()

    pivot = pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)

    return pivot


def generate_opportunities_df(time_df: pd.DataFrame, weeks_back: int) -> pd.DataFrame:
    """Generate Opportunities sheet with hours and last activity."""
    # Load full project codes with all columns
    project_codes = load_project_codes_full()

    # Filter time entries to last N weeks (exclude summary rows)
    weeks = get_weeks_back(weeks_back)
    time_df = time_df[
        (time_df["category"] != ">>> WEEK TOTAL") &
        (time_df["week_beginning"].isin(weeks))
    ].copy()

    # Filter to entries with opportunity_id
    time_df = time_df[time_df["opportunity_id"].notna() & (time_df["opportunity_id"] != "")]

    # Aggregate hours by opportunity_id
    opp_hours = time_df.groupby("opportunity_id").agg({
        "hours": "sum",
        "week_beginning": "max"  # Last activity date
    }).reset_index()
    opp_hours = opp_hours.rename(columns={
        "hours": "Hours (12 wks)",
        "week_beginning": "Last Activity"
    })

    # Join with project codes by opportunity_id (exact match)
    merged = project_codes.merge(
        opp_hours,
        left_on="code",
        right_on="opportunity_id",
        how="inner"
    )

    # Filter to opportunities with hours > 0
    merged = merged[merged["Hours (12 wks)"] > 0]

    # Select and order columns for report
    report_columns = [
        'JDA OpptyID',
        'Account Name',
        'Opportunity Name',
        'JDA Industry',
        'Stage',
        'Booking Amount [USD]',
        'Close Date',
        'Next Step',
        'Hours (12 wks)',
        'Last Activity',
    ]

    # Only include columns that exist
    available_columns = [c for c in report_columns if c in merged.columns]
    result = merged[available_columns].copy()

    # Sort by Hours descending
    result = result.sort_values("Hours (12 wks)", ascending=False)

    # Add TOTAL row
    total_row = {col: "" for col in available_columns}
    total_row[available_columns[0]] = "TOTAL"
    if "Hours (12 wks)" in available_columns:
        total_row["Hours (12 wks)"] = result["Hours (12 wks)"].sum()

    result = pd.concat([result, pd.DataFrame([total_row])], ignore_index=True)

    return result


def write_manager_report(weekly_df: pd.DataFrame, opps_df: pd.DataFrame, output_path: Path):
    """Write manager report to Excel with formatting."""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True)
    total_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet A: Weekly Hours
    ws_hours = wb.active
    ws_hours.title = "Weekly Hours"

    for r_idx, row in enumerate(dataframe_to_rows(weekly_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_hours.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

            # Header row
            if r_idx == 1:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # TOTAL row (last row)
            if r_idx == len(weekly_df) + 1:
                cell.fill = total_fill
                cell.font = total_font

    # Adjust column widths
    for col in ws_hours.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_hours.column_dimensions[column].width = max_length + 2

    # Sheet B: Opportunities
    ws_opps = wb.create_sheet("Opportunities")

    for r_idx, row in enumerate(dataframe_to_rows(opps_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_opps.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

            # Header row
            if r_idx == 1:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # TOTAL row (last row)
            if r_idx == len(opps_df) + 1:
                cell.fill = total_fill
                cell.font = total_font

    # Adjust column widths
    for col in ws_opps.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_opps.column_dimensions[column].width = min(max_length + 2, 50)

    wb.save(output_path)


def generate_manager_report(weeks_back: int | None = None):
    """Main entry point: generate manager report Excel file.

    Reads from existing time_entries_preview.xlsx (generated by 'python run.py preview').
    This is fast since it uses pre-processed data with no AI calls.

    Args:
        weeks_back: Number of weeks to include. If None, uses config default.
    """
    settings = get_settings()
    if weeks_back is None:
        weeks_back = settings.get("report", {}).get("weeks_back", 12)

    input_path = Path("data/output/time_entries_preview.xlsx")
    output_path = Path("data/output/manager_report.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Check that preview file exists
    if not input_path.exists():
        print(f"Error: {input_path} not found.")
        print()
        print("Please generate the preview file first:")
        print("  python run.py preview --weeks 12")
        print()
        print("Then review/edit the Excel file, and run this report again.")
        return

    print(f"Generating manager report (last {weeks_back} weeks)...")
    print()

    # Read from existing preview file (fast, no AI)
    print(f"Reading from {input_path}...")
    time_df = pd.read_excel(input_path)

    # Generate Weekly Hours sheet
    print("Building Weekly Hours summary...")
    weekly_df = generate_weekly_hours_df(time_df, weeks_back)

    # Generate Opportunities sheet
    print("Building Opportunities summary...")
    opps_df = generate_opportunities_df(time_df, weeks_back)

    # Write to Excel
    print("Writing Excel report...")
    write_manager_report(weekly_df, opps_df, output_path)

    print()
    print(f"Manager report generated: {output_path}")
    print()
    print(f"  - Weekly Hours: {len(weekly_df) - 1} weeks")
    print(f"  - Opportunities: {len(opps_df) - 1} accounts with hours")
