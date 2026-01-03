"""
Excel writer with formatting, colors and table.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


def write_excel_with_formatting(df: pd.DataFrame, output_path: Path) -> None:
    """Write DataFrame to Excel with colors and table formatting."""
    
    # Save basic Excel first
    df.to_excel(output_path, index=False, sheet_name="Time Entries")
    
    # Load and format
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Define colors
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Original
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Autofilled
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Week total
    
    # Find column indices
    columns = {cell.value: cell.column for cell in ws[1]}
    category_col = columns.get("category")
    autofill_col = columns.get("is_autofilled")
    
    # Apply colors to data rows
    for row_idx in range(2, ws.max_row + 1):
        category_value = ws.cell(row=row_idx, column=category_col).value if category_col else None
        is_autofilled = ws.cell(row=row_idx, column=autofill_col).value if autofill_col else False
        
        # Determine fill color
        if category_value == ">>> WEEK TOTAL":
            fill = red_fill
        elif is_autofilled:
            fill = yellow_fill
        else:
            fill = green_fill
        
        # Apply to all cells in row
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
    
    # Create table
    table_range = f"A1:{chr(64 + ws.max_column)}{ws.max_row}"
    table = Table(displayName="TimeEntries", ref=table_range)
    
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    wb.save(output_path)